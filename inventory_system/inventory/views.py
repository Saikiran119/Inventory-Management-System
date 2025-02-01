# from django.shortcuts import render, redirect, get_object_or_404
# Import necessary models
from datetime import timedelta
from django.contrib.auth import authenticate, login
from django.db.models import Sum, F
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from .forms import ProductForm, SaleForm, UserCreationForm
# Import necessary models
from .models import Product, Sale, InventorySale, Order
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .forms import OrderForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView
# from django.db import transaction


# View to show inventory
@login_required
def inventory_list(request):
    products = Product.objects.all()
    return render(request, 'inventory/inventory_list.html', {'products': products})


def inventory_view(request):
    # Logic to fetch products, sales data, or any inventory-related information
    products = Product.objects.all()  # Example: Fetch all products
    return render(request, 'inventory/inventory_list.html', {'products': products})


class InventoryListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'inventory/inventory_list.html'
    context_object_name = 'products'


# View to add a new product
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('inventory:inventory_list')
    else:
        form = ProductForm()
    return render(request, 'inventory/add_product.html', {'form': form})


# View to edit a product
def edit_product(request, pk):
    product = Product.objects.get(id=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect('inventory:inventory_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'inventory/add_product.html', {'form': form})


# View to delete a product
def delete_product(request, pk):
    # Get the product and related sale
    product = get_object_or_404(Product, pk=pk)

    # Optionally, delete related sales here if necessary
    InventorySale.objects.filter(product=product).delete()  # Use the correct model class

    # Delete the product
    product.delete()

    return redirect('inventory:inventory_list')  # Redirect to the inventory list


# Login view
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        # Authenticate the user
        user = authenticate(request, username=username, password=password)

        if user is not None:
            # Log the user in
            login(request, user)
            return redirect('inventory:inventory_list')  # Redirect to your desired page after login
        else:
            # Add error message to the messages framework
            messages.error(request, 'Invalid username or password')
            return render(request, 'inventory/login.html')  # Render the login page again

    return render(request, 'inventory/login.html')


def signup_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Account created successfully! You can now log in.')
            return redirect('inventory:login')  # Redirect to login page after successful sign-up
        else:
            messages.error(request, 'Error creating account. Please try again.')
    else:
        form = UserCreationForm()

    return render(request, 'inventory/signup.html', {'form': form})


def signup(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Account created successfully! You can now log in.')
            return redirect('inventory:login')  # Redirect to login page after successful sign-up
        else:
            messages.error(request, 'Error creating account. Please try again.')
    else:
        form = UserCreationForm()  # Render an empty form for GET request

    return render(request, 'inventory/signup.html', {'form': form})


# View to add a sale and update inventory
def low_stock_alerts(request):
    threshold = 5  # Set your low stock threshold
    low_stock_products = Product.objects.filter(quantity__lte=threshold)
    return render(request, 'inventory/low_stock_alerts.html', {'low_stock_products': low_stock_products})


def sales_summary(request):
    start_date = timezone.now() - timedelta(days=30)

    # Correct query to use the 'date' field from the Sale model, and annotate with total quantity sold
    sales = Sale.objects.filter(date__gte=start_date).values('product__name').annotate(
        total_sales=Sum('quantity')
    ).order_by('product__name')  # Optional: order by product name or any other field as needed

    return render(request, 'inventory/sales_summary.html', {'sales': sales})


def home(request):
    return render(request, 'inventory/home.html')


# Add sale view
def add_sale(request, product_id):
    product = Product.objects.get(id=product_id)

    if request.method == 'POST':
        form = SaleForm(request.POST)
        if form.is_valid():
            sale = form.save(commit=False)
            sale.product = product
            sale.total_price = sale.quantity * product.price  # Calculate total price
            sale.save()
            # Update product quantity after sale
            product.quantity -= sale.quantity
            product.save()

            # Create an order after sale
            order = Order.objects.create(product=product, user=request.user, quantity=sale.quantity)
            return redirect('inventory:inventory_list')  # Or redirect to order success page

    else:
        form = SaleForm()

    return render(request, 'inventory/add_sale.html', {'form': form, 'product': product})


# Delete sale view
def delete_sale(request, sale_id):
    sale = get_object_or_404(InventorySale, id=sale_id)  # Make sure this is 'Sale' model if you're using 'Sale'
    sale.delete()
    return redirect('inventory:inventory_list')  # Redirect to inventory list after deletion


def product_list(request):
    # Retrieve all products from the database
    products = Product.objects.all()

    # Render the product list template
    return render(request, 'inventory/product_list.html', {'products': products})


def reports_view(request):
    # Low-stock products
    low_stock_products = Product.objects.filter(quantity__lt=F('low_stock_threshold'))

    # Sales summary
    sales_data = (
        Sale.objects.values(product_name=F('product__name'))
        .annotate(
            total_quantity=Sum('quantity'),
            total_sales=Sum(F('quantity') * F('product__price'))
        )
    )

    context = {
        'low_stock_products': low_stock_products,
        'sales_data': sales_data,
    }
    return render(request, 'inventory/reports.html', context)


def export_low_stock_excel(request):
    low_stock_products = Product.objects.filter(quantity__lt=F('low_stock_threshold'))
    print("Low Stock Products:", low_stock_products)  # Debugging line

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Low Stock Products"

    headers = ['Product Name', 'Category', 'Quantity', 'Low Stock Threshold']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    for row_num, product in enumerate(low_stock_products, 2):
        sheet[f'A{row_num}'] = product.name
        sheet[f'B{row_num}'] = product.category
        sheet[f'C{row_num}'] = product.quantity
        sheet[f'D{row_num}'] = product.low_stock_threshold

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="low_stock_report.xlsx"'
    wb.save(response)

    return response


def export_sales_summary_excel(request):
    start_date = timezone.now() - timedelta(days=30)

    # Retrieve sales data for the past 30 days
    sales_data = (
        Sale.objects.filter(date__gte=start_date)
        .values('product__name')
        .annotate(total_quantity=Sum('quantity'), total_sales=Sum(F('quantity') * F('product__price')))
        .order_by('product__name')
    )

    # Create an Excel workbook and sheet
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sales Summary"

    # Add headers
    headers = ['Product Name', 'Total Quantity Sold', 'Total Sales']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    # Add data rows
    for row_num, data in enumerate(sales_data, 2):
        sheet[f'A{row_num}'] = data['product__name']
        sheet[f'B{row_num}'] = data['total_quantity']
        sheet[f'C{row_num}'] = data['total_sales']

    # Create an HTTP response with the content type set to Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_summary_report.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response


def export_low_stock_csv(request):
    # Fetch products with quantity below the low stock threshold
    low_stock_products = Product.objects.filter(quantity__lt=F('low_stock_threshold'))

    # Create a CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="low_stock_report.csv"'

    # Write data to CSV
    writer = csv.writer(response)
    writer.writerow(['Product Name', 'Quantity', 'Low Stock Threshold'])  # Removed 'Category'

    # Write rows of product data
    for product in low_stock_products:
        writer.writerow([product.name, product.quantity, product.low_stock_threshold])  # Removed 'Category'

    return response


def export_sales_summary_csv(request):
    # Get the start date for the sales summary (e.g., last 30 days)
    start_date = timezone.now() - timedelta(days=30)

    # Retrieve sales data for the past 30 days, and annotate with total sales and quantity sold
    sales_data = Sale.objects.filter(date__gte=start_date).values('product__name').annotate(
        total_quantity=Sum('quantity'),
        total_sales=Sum(F('quantity') * F('product__price'))
    ).order_by('product__name')  # Optionally, you can order by product name or any other field

    # Create the HTTP response with content type CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sales_summary_report.csv"'

    # Create a CSV writer object
    writer = csv.writer(response)

    # Write the headers for the CSV file
    writer.writerow(['Product Name', 'Total Quantity Sold', 'Total Sales'])

    # Write data rows for each sale in the dataset
    for data in sales_data:
        writer.writerow([data['product__name'], data['total_quantity'], data['total_sales']])

    return response


# Assuming you have an order form that includes quantity
@login_required
def place_order(request, product_id):
    product = Product.objects.get(id=product_id)

    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.product = product
            order.user = request.user
            order.total_price = product.price * order.quantity
            order.status = 'pending'  # Default status when order is first placed
            order.save()

            # Update the stock
            if product.quantity >= order.quantity:
                product.update_quantity(-order.quantity)  # Reduce the stock
            else:
                messages.error(request, "Not enough stock.")
                return redirect('inventory:place_order', product_id=product.id)

            # Update status to 'success' once the order is processed
            order.status = 'success'
            order.save()

            # Redirect to the order success page
            return redirect('inventory:order_success', order_id=order.id)
        else:
            messages.error(request, "There was an error with your order.")
    else:
        form = OrderForm()

    return render(request, 'inventory/place_order.html', {'form': form, 'product': product})


@login_required
def order_success(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if order.status == 'success':
        return render(request, 'inventory/order_success.html', {'order': order})
    else:
        return render(request, 'inventory/order_success.html', {'order': order})


@login_required
def order_failed(request, order_id):
    # Fetch the order by its ID, or return a 404 error if it doesn't exist
    order = get_object_or_404(Order, id=order_id)

    # Check the order status
    if order.status != 'failed':
        # If the order status is not 'failed', redirect to a different page (e.g., order success page or inventory list)
        return redirect('inventory:order_success', order_id=order.id)

    # Render the order_failed.html template
    return render(request, 'inventory/order_failed.html', {'order': order})


@login_required
def order_history(request, product_id=None):
    if product_id:
        # If a product_id is passed, filter orders by product
        orders = Order.objects.filter(product_id=product_id).order_by('-order_date')
    else:
        # If no product_id is passed, display all orders
        orders = Order.objects.all().order_by('-order_date')

    return render(request, 'inventory/order_history.html', {
        'orders': orders,
        'product_id': product_id,  # Pass product_id to the template to show the filter
    })


def mark_order_as_success(request, order_id):
    order = Order.objects.get(id=order_id)
    order.mark_as_success()  # This will change the status to 'success'
    return redirect('inventory:order_success', order_id=order.id)


def order_confirmation(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    return render(request, 'inventory/order_confirmation.html', {'order': order})